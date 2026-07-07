from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
README = ROOT / "README.md"
out_path = ROOT / "Catalogue" / "integration-details.xlsx"

headers = ["Flow Name", "Entity", "Info Flow Name", "Source System", "Target System",
           "Source Connector", "Target Connector", "Integration Pattern", "Complexity"]

wb = Workbook()
wb.remove(wb.active)

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="1a3c6e", end_color="1a3c6e", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

text = README.read_text(encoding="utf-8")
lines = text.split("\n")

INT_SEP = "|------|--------|-----------|--------|--------|-------------|-------------|---------|------------|"

sections_data = {}
current_section = None
current_rows = []
in_integration = False

for line in lines:
    stripped = line.strip()

    if stripped.startswith("# ") and not stripped.startswith("## ") and "Higher Education Process" not in stripped:
        if current_section and current_rows:
            sections_data[current_section] = list(current_rows)
        # Extract clean section name from H1 (remove codes)
        m = re.match(r"#\s*(?:\d+:\s*)?(.+?)\s*-", stripped)
        if m:
            current_section = m.group(1).strip()
        else:
            current_section = stripped.lstrip("# ").strip()[:31]
        current_rows = []
        in_integration = False

    if stripped == "#### Integration Details":
        in_integration = True
        continue

    if in_integration:
        if not stripped:
            continue
        if stripped.startswith("| Flow") or stripped.startswith("|------"):
            continue
        if stripped.startswith("|"):
            parts = [p.strip() for p in stripped.split("|")[1:-1]]
            if len(parts) == 9:
                current_rows.append(tuple(parts))
        else:
            in_integration = False

if current_section and current_rows:
    sections_data[current_section] = list(current_rows)

total_rows = 0
for sheet_title, rows in sections_data.items():
    ws = wb.create_sheet(title=sheet_title[:31])
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Calibri", size=9)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 14

    ws.auto_filter.ref = f"A1:I{len(rows)+1}"
    ws.freeze_panes = "A2"
    total_rows += len(rows)

wb.save(str(out_path))
print(f"XLSX written: {out_path}")
print(f"Sheets: {len(sections_data)}, total integration rows: {total_rows}")
