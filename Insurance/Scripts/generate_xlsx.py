"""Generate Insurance XLSX catalogue from README.md integration tables."""

from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent
README = ROOT / "README.md"
out_path = ROOT / "Catalogue" / "integration-details.xlsx"

headers = ["Flow Name", "Entity", "Info Flow Name", "Source System", "Target System",
           "Source Connector", "Target Connector", "Integration Pattern", "Complexity", "Complexity Reason"]

REASON_MAP = {
    ("High", "API-led (Real-time)"):   "Real-time transaction boundary with strict consistency, auditing, and zero-data-loss requirements",
    ("High", "Event-driven"):         "Mission-critical event processing requiring guaranteed delivery, sequencing, and audit trail",
    ("High", "Batch (Real-time)"):    "High-frequency near-real-time batch processing with minimal latency tolerance",
    ("High", "Batch (Scheduled)"):    "Large-volume scheduled processing with data integrity validation and reconciliation",
    ("High", "Batch (ETL)"):          "Large-scale ETL pipeline requiring complex transformations, deduplication, and reconciliation",
    ("High", "Streaming (Real-time)"): "Continuous streaming data pipeline with exactly-once semantics and failover requirements",
    ("Medium", "API-led (Real-time)"): "API integration requiring moderate transformation, error handling, and data reconciliation",
    ("Medium", "Event-driven"):       "Event-driven integration with intermediate complexity in event routing and state management",
    ("Medium", "Batch (Real-time)"):  "Periodic near-real-time batch with data validation and transformation needs",
    ("Medium", "Batch (Scheduled)"):  "Scheduled batch transfer requiring field mapping, validation, and exception handling",
    ("Medium", "Batch (ETL)"):        "Intermediate ETL process with data cleansing, enrichment, and staging requirements",
    ("Simple", "API-led (Real-time)"): "Direct API integration with minimal transformation and single-system lookup",
    ("Simple", "Event-driven"):       "Simple event notification with fire-and-forget delivery semantics",
    ("Simple", "Batch (Real-time)"):  "Minimal near-real-time batch with basic data pass-through",
    ("Simple", "Batch (Scheduled)"):  "Straightforward periodic data transfer with no real-time constraints",
    ("Simple", "Batch (ETL)"):        "Basic ETL pipeline with direct mapping and flat-file exchange",
}

wb = Workbook()
wb.remove(wb.active)

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="1a3c6e", end_color="1a3c6e", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

text = README.read_text(encoding="utf-8")
lines = text.split("\n")

sections_data = {}
current_section = None
current_rows = []
in_integration = False

for line in lines:
    stripped = line.strip()

    if stripped.startswith("# ") and not stripped.startswith("## ") and "Insurance" not in stripped and "Table" not in stripped:
        if current_section and current_rows:
            sections_data[current_section] = list(current_rows)
        m = re.match(r"\d+: ([A-Za-z &,]+)", stripped.lstrip("# "))
        current_section = m.group(1).strip() if m else stripped.lstrip("# ").strip()[:31]
        current_rows = []
        in_integration = False

    if stripped == "#### Integration Details":
        in_integration = True
        continue

    if in_integration:
        if not stripped:
            continue
        if stripped.startswith("| Flow") or stripped.startswith("|------"):
            continue
        if stripped.startswith("|"):
            parts = [p.strip() for p in stripped.split("|")[1:-1]]
            if len(parts) in (9, 10):
                current_rows.append(tuple(parts))
        else:
            in_integration = False

if current_section and current_rows:
    sections_data[current_section] = list(current_rows)

total_rows = 0

for sec_name, rows in sections_data.items():
    ws = wb.create_sheet(title=sec_name[:31])

    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        complexity = row_data[8] if len(row_data) == 10 else ""
        pattern = row_data[7] if len(row_data) == 10 else ""
        complexity_reason = REASON_MAP.get((complexity, pattern), "")

        for col_idx, value in enumerate(row_data[:9], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        cell = ws.cell(row=row_idx, column=10, value=complexity_reason)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 55

    total_rows += len(rows)

out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)

print(f"XLSX written: {out_path}")
print(f"  Sections: {len(sections_data)}")
print(f"  Total integration rows: {total_rows}")
print()
